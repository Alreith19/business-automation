import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime

# Define paths
input_folder = 'D:/Business 2024/Programs/BusinessAutomation/InputFiles/'
output_folder = 'D:/Business 2024/Programs/BusinessAutomation/OutputFiles/'

# Read the CSV files
try:
    labor_hours_df = pd.read_csv(os.path.join(input_folder, 'labor_hours.csv'))
    material_purchases_df = pd.read_csv(os.path.join(input_folder, 'material_purchases.csv'))
except FileNotFoundError as e:
    print(f"Error: {e}")
    exit(1)

# Print headers of the DataFrames for debugging
print("Labor Hours DataFrame Columns:", labor_hours_df.columns.tolist())
print("Material Purchases DataFrame Columns:", material_purchases_df.columns.tolist())

# Check if required columns are present in labor_hours_df
required_columns_labor = ['Employee', 'Date', 'Scope', 'Hours', 'Rate']
for col in required_columns_labor:
    if col not in labor_hours_df.columns:
        print(f"Error: Missing column '{col}' in labor_hours.csv")
        exit(1)

# Check if required columns are present in material_purchases_df
required_columns_material = ['Date', 'Item', 'Scope', 'Cost']
for col in required_columns_material:
    if col not in material_purchases_df.columns:
        print(f"Error: Missing column '{col}' in material_purchases.csv")
        exit(1)


# Function to calculate labor costs including overhead
def calculate_labor_costs(df):
    df['Cost'] = df['Rate'] * df['Hours']
    df['Cost_with_tax'] = df['Cost'] * 1.07  # Example tax rate of 7%
    df['Cost_with_overhead'] = df['Cost_with_tax'] * 1.15  # Example overhead of 15%
    return df


# Function to calculate material costs including overhead
def calculate_material_costs(df):
    df['Cost_with_tax'] = df['Cost'] * 1.07  # Example tax rate of 7%
    df['Cost_with_overhead'] = df['Cost_with_tax'] * 1.15  # Example overhead of 15%
    return df


# Process labor hours
labor_hours_df = calculate_labor_costs(labor_hours_df)

# Summarize labor by scope
labor_summary = labor_hours_df.groupby('Scope').agg({
    'Hours': 'sum',
    'Cost': 'sum',
    'Cost_with_tax': 'sum',
    'Cost_with_overhead': 'sum'
}).reset_index()

# Process material purchases
material_purchases_df = calculate_material_costs(material_purchases_df)

# Summarize materials by scope
material_summary = material_purchases_df.groupby('Scope').agg({
    'Cost': 'sum',
    'Cost_with_tax': 'sum',
    'Cost_with_overhead': 'sum'
}).reset_index()

# Calculate total costs for the customer according to scope
total_costs_df = labor_summary[['Scope', 'Cost_with_overhead']].rename(
    columns={'Cost_with_overhead': 'Labor_Cost_with_Overhead'})
total_costs_df = total_costs_df.merge(material_summary[['Scope', 'Cost_with_overhead']], on='Scope',
                                      how='outer').rename(columns={'Cost_with_overhead': 'Material_Cost_with_Overhead'})
total_costs_df['Total_Cost'] = total_costs_df['Labor_Cost_with_Overhead'].fillna(0) + total_costs_df[
    'Material_Cost_with_Overhead'].fillna(0)

# Create Excel report
date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
output_path = os.path.join(output_folder, f"{date}_detailed_summary_report_v2.xlsx")
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    # Write summaries to separate sheets
    labor_summary.to_excel(writer, sheet_name='Labor Summary', index=False)
    material_summary.to_excel(writer, sheet_name='Material Summary', index=False)

    # Write total costs summary
    total_costs_df.to_excel(writer, sheet_name='Total Costs by Scope', index=False)

# Autofit columns and add formatting
try:
    wb = load_workbook(output_path)
    for sheet in wb.sheetnames:
        worksheet = wb[sheet]
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    print(f"Error calculating length for cell {cell.coordinate}: {e}")
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width

        # Apply currency format to specific columns
        if sheet in ['Labor Summary', 'Material Summary', 'Total Costs by Scope']:
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2,
                                           max_col=worksheet.max_column):
                for cell in row:
                    cell.number_format = '$#,##0.00'

    wb.save(output_path)
except Exception as e:
    print(f"Error processing the Excel file: {e}")

print(f"Detailed summary report has been generated: {output_path}")
